from flask import Flask, render_template, jsonify, request, send_from_directory, make_response, redirect, url_for, session
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import json
import os
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
import logging
import io
import base64
import pickle
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import BarChart, LineChart, Reference
import tempfile
import sqlite3
import requests
import secrets

from functools import wraps

app = Flask(__name__)
app.secret_key = secrets.token_hex(16)  # Generate a secure secret key
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['DOWNLOAD_FOLDER'] = 'downloads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Global variable to store the processed data
data_cache = {}
raw_data = None

# Database initialization
def init_database():
    """Initialize the user database"""
    conn = sqlite3.connect('users.db')
    cursor = conn.cursor()
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            email TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            full_name TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'user',
            organization TEXT,
            phone TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            last_login TIMESTAMP,
            is_active BOOLEAN DEFAULT 1
        )
    ''')
    
    # Create default admin user if it doesn't exist
    cursor.execute('SELECT COUNT(*) FROM users WHERE role = "admin"')
    if cursor.fetchone()[0] == 0:
        admin_password_hash = generate_password_hash('admin123')
        cursor.execute('''
            INSERT INTO users (username, email, password_hash, full_name, role, organization)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', ('admin', 'admin@healthdashboard.com', admin_password_hash, 'System Administrator', 'admin', 'Health Department'))
    
    conn.commit()
    conn.close()



#############################################################################################################################
from flask import Flask, request, render_template, redirect, url_for, flash
# Create DB and table if not exists
def init_db():
    conn = sqlite3.connect('alerts.db')
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS alerts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            reporter_name TEXT NOT NULL,
            email TEXT NOT NULL,
            phone TEXT NOT NULL,
            location TEXT NOT NULL,
            hospital_email TEXT,
            hospital_location TEXT,
            disease TEXT NOT NULL,
            cases INTEGER NOT NULL,
            details TEXT
        )
    ''')
    conn.commit()
    conn.close()

init_db()

# @app.route('/submit_alert', methods=['POST'])
# def submit_alert():
#     try:
#         reporter_name = request.form['alertName']
#         email = request.form['alertEmail']
#         phone = request.form['alertPhone']
#         location = request.form['alertLocation']
#         hospital_email = request.form.get('hospitalEmail', '')
#         hospital_location = request.form.get('hospitalLocation', '')
#         disease = request.form['alertDisease']
#         cases = int(request.form['alertCases'])
#         details = request.form.get('alertDetails', '')

#         conn = sqlite3.connect('alerts.db')
#         cursor = conn.cursor()
#         cursor.execute('''
#             INSERT INTO alerts 
#             (reporter_name, email, phone, location, hospital_email, hospital_location, disease, cases, details)
#             VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
#         ''', (reporter_name, email, phone, location, hospital_email, hospital_location, disease, cases, details))
#         conn.commit()
#         conn.close()
#         print("WEll done manu")

#         flash("Alert submitted successfully!", "success")
#         return redirect(url_for('index'))
#     except Exception as e:
#         flash(f"Error: {str(e)}", "danger")
#         return redirect(url_for('index'))
################################################################################################################################
# Authentication decorator
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'error': 'Authentication required'}), 401
        return f(*args, **kwargs)
    return decorated_function

def role_required(required_roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                return jsonify({'error': 'Authentication required'}), 401
            
            user_role = session.get('user_role')
            if user_role not in required_roles:
                return jsonify({'error': 'Insufficient permissions'}), 403
            
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# User management functions
def create_user(username, email, password, full_name, role='user', organization='', phone=''):
    """Create a new user"""
    try:
        conn = sqlite3.connect('users.db')
        cursor = conn.cursor()
        
        password_hash = generate_password_hash(password)
        
        cursor.execute('''
            INSERT INTO users (username, email, password_hash, full_name, role, organization, phone)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (username, email, password_hash, full_name, role, organization, phone))
        
        conn.commit()
        user_id = cursor.lastrowid
        conn.close()
        
        return user_id
    except sqlite3.IntegrityError as e:
        if 'username' in str(e):
            raise ValueError('Username already exists')
        elif 'email' in str(e):
            raise ValueError('Email already exists')
        else:
            raise ValueError('User creation failed')

def authenticate_user(username, password):
    """Authenticate user and return user info"""
    conn = sqlite3.connect('users.db')
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT id, username, email, password_hash, full_name, role, organization, phone, is_active
        FROM users WHERE username = ? OR email = ?
    ''', (username, username))
    
    user = cursor.fetchone()
    
    if user and check_password_hash(user[3], password) and user[8]:  # user[8] is is_active
        # Update last login
        cursor.execute('UPDATE users SET last_login = CURRENT_TIMESTAMP WHERE id = ?', (user[0],))
        conn.commit()
        
        user_info = {
            'id': user[0],
            'username': user[1],
            'email': user[2],
            'full_name': user[4],
            'role': user[5],
            'organization': user[6],
            'phone': user[7]
        }
        
        conn.close()
        return user_info
    
    conn.close()
    return None

# Authentication routes
@app.route('/')
def index():
    """Serve login page or dashboard based on authentication"""
    if 'user_id' not in session:
        return render_template('dashboard.html')
    return render_template('login.html')

@app.route('/login')
def login_page():
    """Serve login page"""
    redirect_tab = request.args.get('redirect', 'overview')
    return render_template('login.html')

@app.route('/register')
def register_page():
    """Serve registration page"""
    return render_template('register.html')

@app.route('/dashboard')

def dashboard():
    """Serve the main dashboard (requires authentication)"""
    return render_template('dashboard.html')

@app.route('/api/auth/register', methods=['POST'])
def api_register():
    """User registration API"""
    try:
        data = request.get_json()
        
        required_fields = ['username', 'email', 'password', 'full_name', 'role']
        for field in required_fields:
            if field not in data or not data[field].strip():
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        # Validate role
        if data['role'] not in ['user', 'health_supervisor']:
            return jsonify({'error': 'Invalid role'}), 400
        
        # Validate email format (basic)
        if '@' not in data['email'] or '.' not in data['email']:
            return jsonify({'error': 'Invalid email format'}), 400
        
        # Validate password strength
        if len(data['password']) < 6:
            return jsonify({'error': 'Password must be at least 6 characters'}), 400
        
        user_id = create_user(
            username=data['username'].strip(),
            email=data['email'].strip().lower(),
            password=data['password'],
            full_name=data['full_name'].strip(),
            role=data['role'],
            organization=data.get('organization', '').strip(),
            phone=data.get('phone', '').strip()
        )
        
        return jsonify({
            'success': True,
            'message': 'Registration successful! Please login.',
            'user_id': user_id
        })
        
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        logger.error(f"Registration error: {str(e)}")
        return jsonify({'error': 'Registration failed'}), 500

@app.route('/api/auth/login', methods=['POST'])
def api_login():
    """User login API"""
    try:
        data = request.get_json()
        
        if not data.get('username') or not data.get('password'):
            return jsonify({'error': 'Username and password required'}), 400
        
        user = authenticate_user(data['username'], data['password'])
        
        if user:
            # Create session
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['user_role'] = user['role']
            session['full_name'] = user['full_name']
            
            return jsonify({
                'success': True,
                'message': 'Login successful',
                'user': {
                    'username': user['username'],
                    'full_name': user['full_name'],
                    'role': user['role'],
                    'organization': user['organization']
                }
            })
        else:
            return jsonify({'error': 'Invalid username or password'}), 401
            
    except Exception as e:
        logger.error(f"Login error: {str(e)}")
        return jsonify({'error': 'Login failed'}), 500

@app.route('/api/auth/logout', methods=['POST'])
def api_logout():
    """User logout API"""
    session.clear()
    return jsonify({'success': True, 'message': 'Logged out successfully'})

@app.route('/api/auth/profile')

def api_profile():
    """Get user profile information"""
    try:
        conn = sqlite3.connect('users.db')
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT username, email, full_name, role, organization, phone, created_at, last_login
            FROM users WHERE id = ?
        ''', (session['user_id'],))
        
        user = cursor.fetchone()
        conn.close()
        
        if user:
            return jsonify({
                'username': user[0],
                'email': user[1],
                'full_name': user[2],
                'role': user[3],
                'organization': user[4],
                'phone': user[5],
                'created_at': user[6],
                'last_login': user[7]
            })
        else:
            return jsonify({'error': 'User not found'}), 404
            
    except Exception as e:
        logger.error(f"Profile error: {str(e)}")
        return jsonify({'error': 'Failed to fetch profile'}), 500




def load_and_process_data(file_path='Final_data.csv'):
    """Load and process the CSV data"""
    global raw_data, data_cache
    
    try:
        # Load the data
        logger.info(f"Loading data from {file_path}")
        raw_data = pd.read_csv(file_path)
        
        # Print column names for debugging
        logger.info(f"CSV columns: {list(raw_data.columns)}")
        logger.info(f"Data shape: {raw_data.shape}")
        
        # Data Cleaning and Preparation
        raw_data['Cases'] = pd.to_numeric(raw_data['Cases'], errors='coerce').fillna(0).astype(int)
        raw_data['Deaths'] = pd.to_numeric(raw_data['Deaths'], errors='coerce').fillna(0).astype(int)
        
        # Ensure year/mon/day exist and build date column robustly
        if all(col in raw_data.columns for col in ['year', 'mon', 'day']):
            try:
                raw_data['date'] = pd.to_datetime(
                    raw_data[['year', 'mon', 'day']].rename(columns={'mon':'month'}).astype(int),
                    errors='coerce'
                )
            except Exception:
                raw_data['date'] = pd.to_datetime(raw_data[['year', 'mon', 'day']].astype(str).agg('-'.join, axis=1), errors='coerce')
        elif 'date' in raw_data.columns:
            raw_data['date'] = pd.to_datetime(raw_data['date'], errors='coerce')
        else:
            raw_data['date'] = pd.NaT
        
        # Fill NaN values
        if 'Temp' in raw_data.columns:
            raw_data['Temp'] = raw_data['Temp'].fillna(raw_data['Temp'].mean())
        else:
            raw_data['Temp'] = np.nan
        if 'preci' in raw_data.columns:
            raw_data['preci'] = raw_data['preci'].fillna(0)
        else:
            raw_data['preci'] = 0
            
        # Ensure state and city columns exist and are strings
        if 'state_ut' not in raw_data.columns:
            raw_data['state_ut'] = 'Unknown'
        else:
            raw_data['state_ut'] = raw_data['state_ut'].fillna('Unknown').astype(str)
        if 'city' not in raw_data.columns:
            raw_data['city'] = 'Unknown'
        else:
            raw_data['city'] = raw_data['city'].fillna('Unknown').astype(str)
        
        # Detect disease column - be more flexible in detection
        possible_disease_cols = ['disease', 'disease_type', 'disease_name', 'Disease', 'illness', 'diagnosis', 'Disease_Type']
        disease_col = None
        
        # First, try exact matches
        for col in raw_data.columns:
            if col in possible_disease_cols:
                disease_col = col
                break
        
        # If no exact match, try case-insensitive matches
        if not disease_col:
            for col in raw_data.columns:
                if col.lower() in [n.lower() for n in possible_disease_cols]:
                    disease_col = col
                    break
        
        # If still no match, look for columns containing 'disease' keyword
        if not disease_col:
            for col in raw_data.columns:
                if 'disease' in col.lower():
                    disease_col = col
                    break
        
        logger.info(f"Disease column detected: {disease_col}")
        
        # Create a mapping from state -> list of cities (unique, sorted)
        city_map = {}
        try:
            for state_name, group in raw_data.groupby('state_ut'):
                cities_for_state = sorted(group['city'].dropna().unique().tolist())
                city_map[state_name] = cities_for_state
        except Exception as e:
            logger.warning(f"Error building city_map: {e}")
            city_map = {}
        
        # Get unique diseases from the actual data
        diseases = []
        if disease_col and disease_col in raw_data.columns:
            # Clean the disease data
            raw_data[disease_col] = raw_data[disease_col].fillna('Unknown').astype(str)
            # Get unique diseases, excluding empty/null values
            unique_diseases = raw_data[disease_col].dropna().unique()
            diseases = sorted([d for d in unique_diseases if d and d.strip() and d.lower() != 'unknown'])
            logger.info(f"Found {len(diseases)} unique diseases: {diseases}")
        
        data_cache = {
            'total_records': int(len(raw_data)),
            'date_range': {
                'min': raw_data['date'].min().strftime('%Y-%m-%d') if pd.notna(raw_data['date'].min()) else None,
                'max': raw_data['date'].max().strftime('%Y-%m-%d') if pd.notna(raw_data['date'].max()) else None
            },
            'states': sorted(raw_data['state_ut'].dropna().unique().tolist()),
            'cities': sorted(raw_data['city'].dropna().unique().tolist()),
            'years': sorted(raw_data['year'].dropna().unique().tolist()) if 'year' in raw_data.columns else [],
            'diseases': diseases,
            'disease_column': disease_col,
            'city_map': city_map
        }
        
        logger.info(f"Data loaded successfully: {len(raw_data)} records, {len(diseases)} diseases found")
        return True
        
    except Exception as e:
        logger.error(f"Error loading data: {str(e)}")
        return False

def get_filtered_data(state=None, city=None, year=None, disease=None, start_date=None, end_date=None):
    """Filter data based on parameters"""
    if raw_data is None:
        return pd.DataFrame()
    
    filtered_data = raw_data.copy()
    
    if state and state != 'all':
        filtered_data = filtered_data[filtered_data['state_ut'] == state]
    
    if city and city != 'all' and 'city' in filtered_data.columns:
        filtered_data = filtered_data[filtered_data['city'] == city]
    
    if year and year != 'all':
        try:
            filtered_data = filtered_data[filtered_data['year'] == int(year)]
        except Exception:
            pass
    
    if disease and disease != 'all' and data_cache.get('disease_column'):
        disease_col = data_cache['disease_column']
        filtered_data = filtered_data[filtered_data[disease_col] == disease]
    
    if start_date:
        filtered_data = filtered_data[filtered_data['date'] >= pd.to_datetime(start_date)]
    
    if end_date:
        filtered_data = filtered_data[filtered_data['date'] <= pd.to_datetime(end_date)]
    
    return filtered_data


# @app.route('/')
# def index():
#     """Serve the main dashboard"""
#     return render_template('dashboard.html')

@app.route('/api/data/overview')

def api_overview():
    """Get overview statistics"""
    try:
        state = request.args.get('state', 'all')
        city = request.args.get('city', 'all')
        year = request.args.get('year', 'all')
        disease = request.args.get('disease', 'all')
        
        filtered_data = get_filtered_data(state, city, year, disease)
        
        if filtered_data.empty:
            return jsonify({
                'total_cases': 0,
                'total_deaths': 0,
                'affected_states': 0,
                'avg_temp': 0,
                'data_points': 0
            })
        
        # Calculate statistics
        total_cases = int(filtered_data['Cases'].sum())
        total_deaths = int(filtered_data['Deaths'].sum())
        affected_states = len(filtered_data['state_ut'].unique())
        avg_temp = round(float(filtered_data['Temp'].mean()), 1) if 'Temp' in filtered_data.columns else 0
        
        return jsonify({
            'total_cases': total_cases,
            'total_deaths': total_deaths,
            'affected_states': affected_states,
            'avg_temp': avg_temp,
            'data_points': len(filtered_data)
        })
        
    except Exception as e:
        logger.error(f"Error in overview API: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/data/timeseries')

def api_timeseries():
    """Get time series data for charts"""
    try:
        state = request.args.get('state', 'all')
        city = request.args.get('city', 'all')
        year = request.args.get('year', 'all')
        disease = request.args.get('disease', 'all')
        
        filtered_data = get_filtered_data(state, city, year, disease)
        
        if filtered_data.empty:
            return jsonify({'cases': [], 'deaths': [], 'labels': []})
        
        # Group by date and aggregate
        time_series = filtered_data.groupby('date').agg({
            'Cases': 'sum',
            'Deaths': 'sum',
            'Temp': 'mean',
            'preci': 'mean'
        }).reset_index()
        
        time_series = time_series.sort_values('date')
        
        # Calculate moving averages
        window_size = min(7, len(time_series))
        if window_size > 1:
            time_series['Cases_Smooth'] = time_series['Cases'].rolling(window=window_size, center=True).mean()
            time_series['Deaths_Smooth'] = time_series['Deaths'].rolling(window=window_size, center=True).mean()
        else:
            time_series['Cases_Smooth'] = time_series['Cases']
            time_series['Deaths_Smooth'] = time_series['Deaths']
        
        time_series['Cases_Smooth'] = time_series['Cases_Smooth'].fillna(time_series['Cases'])
        time_series['Deaths_Smooth'] = time_series['Deaths_Smooth'].fillna(time_series['Deaths'])
        
        return jsonify({
            'cases': time_series['Cases_Smooth'].astype(float).tolist(),
            'deaths': time_series['Deaths_Smooth'].astype(float).tolist(),
            'temperature': time_series['Temp'].astype(float).tolist(),
            'precipitation': time_series['preci'].astype(float).tolist(),
            'labels': time_series['date'].dt.strftime('%Y-%m-%d').tolist()
        })
        
    except Exception as e:
        logger.error(f"Error in timeseries API: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/data/disease_breakdown')
def api_disease_breakdown():
    """Returns aggregated disease counts for filtered data"""
    try:
        state = request.args.get('state', 'all')
        city = request.args.get('city', 'all')
        year = request.args.get('year', 'all')
        
        filtered = get_filtered_data(state=state, city=city, year=year)
        if filtered.empty:
            return jsonify({'labels': [], 'values': [], 'data_points': 0})
        
        disease_col = data_cache.get('disease_column')
        
        # If we have a disease column in the data, use it
        if disease_col and disease_col in filtered.columns:
            # Group by disease and sum cases
            agg = filtered.groupby(disease_col).agg({'Cases': 'sum'}).reset_index()
            
            # Remove any null/unknown diseases and sort by cases
            agg = agg[agg[disease_col].notna() & (agg[disease_col] != 'Unknown')]
            agg = agg.sort_values('Cases', ascending=False)
            
            if len(agg) > 0:
                labels = agg[disease_col].astype(str).tolist()
                values = agg['Cases'].astype(int).tolist()
                
                logger.info(f"Disease breakdown - Found {len(labels)} diseases with data")
                return jsonify({'labels': labels, 'values': values, 'data_points': len(filtered)})
        
        # Fallback: if no disease column or no valid data, return empty
        logger.warning("No valid disease data found, returning empty breakdown")
        return jsonify({'labels': [], 'values': [], 'data_points': len(filtered)})
        
    except Exception as e:
        logger.error(f"Error in disease_breakdown API: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/reports/generate', methods=['POST'])
@login_required
def generate_report():
    """Generate and download report in Excel or image format"""
    try:
        data = request.get_json()
        report_type = data.get('type', 'excel')  # 'excel' or 'image'
        filters = data.get('filters', {})
        
        # Get filtered data
        filtered_data = get_filtered_data(
            state=filters.get('state', 'all'),
            city=filters.get('city', 'all'),
            year=filters.get('year', 'all'),
            disease=filters.get('disease', 'all')
        )
        
        if filtered_data.empty:
            return jsonify({'error': 'No data available for the selected filters'}), 400
        
        # Create downloads directory
        os.makedirs(app.config['DOWNLOAD_FOLDER'], exist_ok=True)
        
        if report_type == 'excel':
            return generate_excel_report(filtered_data, filters)
        elif report_type == 'image':
            return generate_image_report(filtered_data, filters)
        else:
            return jsonify({'error': 'Invalid report type'}), 400
            
    except Exception as e:
        logger.error(f"Error generating report: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.errorhandler(404)
def not_found(error):
    return jsonify({'error': 'Not found'}), 404

@app.errorhandler(500)
def internal_error(error):
    return jsonify({'error': 'Internal server error'}), 500


def generate_excel_report(filtered_data, filters):
    """Generate Excel report"""
    try:
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Disease Analysis Report"
        
        # Report header
        ws['A1'] = 'Disease Analysis Report'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:E1')
        
        # Filter information
        row = 3
        ws[f'A{row}'] = 'Report Filters:'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        for key, value in filters.items():
            if value and value != 'all':
                ws[f'A{row}'] = f'{key.title()}: {value}'
                row += 1
        
        ws[f'A{row}'] = f'Generated on: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        row += 2
        
        # Summary statistics
        ws[f'A{row}'] = 'Summary Statistics'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        total_cases = int(filtered_data['Cases'].sum())
        total_deaths = int(filtered_data['Deaths'].sum())
        affected_states = len(filtered_data['state_ut'].unique())
        
        stats = [
            ['Total Cases', total_cases],
            ['Total Deaths', total_deaths],
            ['Affected States', affected_states],
            ['Data Points', len(filtered_data)]
        ]
        
        for stat in stats:
            ws[f'A{row}'] = stat[0]
            ws[f'B{row}'] = stat[1]
            row += 1
        
        row += 1
        
        # State-wise breakdown
        ws[f'A{row}'] = 'State-wise Breakdown'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        state_data = filtered_data.groupby('state_ut').agg({
            'Cases': 'sum',
            'Deaths': 'sum'
        }).reset_index().sort_values('Cases', ascending=False)
        
        headers = ['State', 'Cases', 'Deaths']
        for i, header in enumerate(headers):
            ws.cell(row=row, column=i+1, value=header).font = Font(bold=True)
        row += 1
        
        for _, state_row in state_data.iterrows():
            ws[f'A{row}'] = state_row['state_ut']
            ws[f'B{row}'] = int(state_row['Cases'])
            ws[f'C{row}'] = int(state_row['Deaths'])
            row += 1
        
        row += 1
        
        # Disease-wise breakdown (if disease column exists)
        if data_cache.get('disease_column') and data_cache['disease_column'] in filtered_data.columns:
            disease_col = data_cache['disease_column']
            ws[f'A{row}'] = 'Disease-wise Breakdown'
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            disease_data = filtered_data.groupby(disease_col).agg({
                'Cases': 'sum',
                'Deaths': 'sum'
            }).reset_index().sort_values('Cases', ascending=False)
            
            headers = ['Disease', 'Cases', 'Deaths']
            for i, header in enumerate(headers):
                ws.cell(row=row, column=i+1, value=header).font = Font(bold=True)
            row += 1
            
            for _, disease_row in disease_data.iterrows():
                ws[f'A{row}'] = disease_row[disease_col]
                ws[f'B{row}'] = int(disease_row['Cases'])
                ws[f'C{row}'] = int(disease_row['Deaths'])
                row += 1
        
        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        # Read file content
        with open(temp_file.name, 'rb') as f:
            file_content = f.read()
        
        # Clean up
        os.unlink(temp_file.name)
        
        # Return base64 encoded file
        file_b64 = base64.b64encode(file_content).decode('utf-8')
        filename = f"disease_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return jsonify({
            'success': True,
            'file_content': file_b64,
            'filename': filename,
            'content_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })
        
    except Exception as e:
        logger.error(f"Error generating Excel report: {str(e)}")
        return jsonify({'error': str(e)}), 500

def generate_image_report(filtered_data, filters):
    """Generate image report with charts"""
    try:
        # Set up the plot
        plt.style.use('seaborn-v0_8')
        fig, axes = plt.subplots(2, 2, figsize=(15, 12))
        fig.suptitle('Disease Analysis Report', fontsize=16, fontweight='bold')
        
        # 1. Cases by State (Top 10)
        state_data = filtered_data.groupby('state_ut')['Cases'].sum().sort_values(ascending=False).head(10)
        axes[0, 0].bar(range(len(state_data)), state_data.values, color='steelblue')
        axes[0, 0].set_title('Top 10 States by Cases')
        axes[0, 0].set_xlabel('States')
        axes[0, 0].set_ylabel('Cases')
        axes[0, 0].set_xticks(range(len(state_data)))
        axes[0, 0].set_xticklabels(state_data.index, rotation=45, ha='right')
        
        # 2. Time series if date column exists
        if 'date' in filtered_data.columns and pd.notna(filtered_data['date']).any():
            time_data = filtered_data.groupby('date')['Cases'].sum().sort_index()
            axes[0, 1].plot(time_data.index, time_data.values, color='red', linewidth=2)
            axes[0, 1].set_title('Cases Over Time')
            axes[0, 1].set_xlabel('Date')
            axes[0, 1].set_ylabel('Cases')
            axes[0, 1].tick_params(axis='x', rotation=45)
        else:
            axes[0, 1].text(0.5, 0.5, 'No date data available', ha='center', va='center', transform=axes[0, 1].transAxes)
            axes[0, 1].set_title('Cases Over Time (No Data)')
        
        # 3. Disease distribution (if available)
        if data_cache.get('disease_column') and data_cache['disease_column'] in filtered_data.columns:
            disease_col = data_cache['disease_column']
            disease_data = filtered_data.groupby(disease_col)['Cases'].sum().sort_values(ascending=False)
            
            colors = plt.cm.Set3(np.linspace(0, 1, len(disease_data)))
            axes[1, 0].pie(disease_data.values, labels=disease_data.index, autopct='%1.1f%%', colors=colors)
            axes[1, 0].set_title('Disease Distribution')
        else:
            axes[1, 0].text(0.5, 0.5, 'No disease data available', ha='center', va='center', transform=axes[1, 0].transAxes)
            axes[1, 0].set_title('Disease Distribution (No Data)')
        
        # 4. Cases vs Deaths scatter
        if len(filtered_data) > 0:
            state_summary = filtered_data.groupby('state_ut').agg({
                'Cases': 'sum',
                'Deaths': 'sum'
            }).reset_index()
            
            axes[1, 1].scatter(state_summary['Cases'], state_summary['Deaths'], alpha=0.7, color='orange')
            axes[1, 1].set_title('Cases vs Deaths by State')
            axes[1, 1].set_xlabel('Cases')
            axes[1, 1].set_ylabel('Deaths')
            
            # Add trend line
            if len(state_summary) > 1:
                z = np.polyfit(state_summary['Cases'], state_summary['Deaths'], 1)
                p = np.poly1d(z)
                axes[1, 1].plot(state_summary['Cases'], p(state_summary['Cases']), "r--", alpha=0.8)
        
        # Add filter information
        filter_text = "Filters: "
        for key, value in filters.items():
            if value and value != 'all':
                filter_text += f"{key.title()}: {value}, "
        filter_text = filter_text.rstrip(', ')
        
        fig.text(0.02, 0.02, filter_text, fontsize=8, ha='left')
        fig.text(0.98, 0.02, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", 
                fontsize=8, ha='right')
        
        plt.tight_layout()
        plt.subplots_adjust(top=0.93, bottom=0.07)
        
        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
        plt.savefig(temp_file.name, dpi=300, bbox_inches='tight')
        plt.close()
        
        # Read file content
        with open(temp_file.name, 'rb') as f:
            file_content = f.read()
        
        # Clean up
        os.unlink(temp_file.name)
        
        # Return base64 encoded file
        file_b64 = base64.b64encode(file_content).decode('utf-8')
        filename = f"disease_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
        
        return jsonify({
            'success': True,
            'file_content': file_b64,
            'filename': filename,
            'content_type': 'image/png'
        })
        
    except Exception as e:
        logger.error(f"Error generating image report: {str(e)}")
        return jsonify({'error': str(e)}), 500

# Keep all existing API endpoints...
@app.route('/api/data/geographic')
def api_geographic():
    """Get geographic distribution data"""
    try:
        state = request.args.get('state', 'all')
        year = request.args.get('year', 'all')
        disease = request.args.get('disease', 'all')
        
        filtered_data = get_filtered_data(state=state, year=year, disease=disease)
        
        if filtered_data.empty or 'Latitude' not in filtered_data.columns or 'Longitude' not in filtered_data.columns:
            return jsonify({'geographic_data': [], '3d_data': {}})
        
        geo_data = filtered_data.groupby(['Latitude', 'Longitude', 'state_ut']).agg({
            'Cases': 'sum',
            'Deaths': 'sum',
            'Temp': 'mean',
            'preci': 'mean'
        }).reset_index()
        
        geo_data = geo_data.dropna(subset=['Latitude', 'Longitude'])
        
        geographic_data = []
        for _, row in geo_data.iterrows():
            geographic_data.append({
                'lat': float(row['Latitude']),
                'lng': float(row['Longitude']),
                'state': str(row['state_ut']),
                'cases': int(row['Cases']),
                'deaths': int(row['Deaths']),
                'temp': float(row['Temp']),
                'precipitation': float(row['preci'])
            })
        
        plot_3d_data = {
            'temperature': geo_data['Temp'].tolist(),
            'precipitation': geo_data['preci'].tolist(),
            'cases': geo_data['Cases'].tolist(),
            'states': geo_data['state_ut'].tolist()
        }
        
        return jsonify({
            'geographic_data': geographic_data,
            '3d_data': plot_3d_data
        })
        
    except Exception as e:
        logger.error(f"Error in geographic API: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/data/trends')
def api_trends():
    """Get trends analysis data"""
    try:
        trend_type = request.args.get('type', 'state')
        entity = request.args.get('entity', 'all')
        disease = request.args.get('disease', 'all')
        
        filtered_data = get_filtered_data(disease=disease)
        
        if trend_type == 'state':
            grouped = filtered_data.groupby('state_ut').agg({
                'Cases': 'sum',
                'Deaths': 'sum'
            }).reset_index()
            
            if entity != 'all':
                grouped = grouped[grouped['state_ut'] == entity]
            
            trend_data = {
                'labels': grouped['state_ut'].tolist(),
                'cases': grouped['Cases'].tolist(),
                'deaths': grouped['Deaths'].tolist()
            }
            
        elif trend_type == 'disease' and data_cache.get('disease_column'):
            disease_col = data_cache['disease_column']
            grouped = filtered_data.groupby(disease_col).agg({
                'Cases': 'sum',
                'Deaths': 'sum'
            }).reset_index()
            
            trend_data = {
                'labels': grouped[disease_col].tolist(),
                'cases': grouped['Cases'].tolist(),
                'deaths': grouped['Deaths'].tolist()
            }
            
        elif trend_type == 'city' and 'city' in filtered_data.columns:
            grouped = filtered_data.groupby('city').agg({
                'Cases': 'sum',
                'Deaths': 'sum'
            }).reset_index()
            
            if entity != 'all':
                grouped = grouped[grouped['city'] == entity]
            
            grouped = grouped.nlargest(10, 'Cases')
            
            trend_data = {
                'labels': grouped['city'].tolist(),
                'cases': grouped['Cases'].tolist(),
                'deaths': grouped['Deaths'].tolist()
            }
            
        elif trend_type == 'year':
            grouped = filtered_data.groupby('year').agg({
                'Cases': 'sum',
                'Deaths': 'sum'
            }).reset_index()
            
            trend_data = {
                'labels': [str(year) for year in grouped['year'].tolist()],
                'cases': grouped['Cases'].tolist(),
                'deaths': grouped['Deaths'].tolist()
            }
        
        else:
            return jsonify({'error': 'Invalid trend type'}), 400
        
        # Calculate growth rates
        cases = trend_data['cases']
        growth_rates = [0]
        for i in range(1, len(cases)):
            if cases[i-1] != 0:
                growth_rate = ((cases[i] - cases[i-1]) / cases[i-1]) * 100
            else:
                growth_rate = 0
            growth_rates.append(round(growth_rate, 2))
        
        trend_data['growth_rates'] = growth_rates
        
        return jsonify(trend_data)
        
    except Exception as e:
        logger.error(f"Error in trends API: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/data/seasonal')
def api_seasonal():
    """Get seasonal analysis data"""
    try:
        disease = request.args.get('disease', 'all')
        filtered_data = get_filtered_data(disease=disease)
        
        if 'mon' in filtered_data.columns:
            filtered_data['month'] = filtered_data['mon']
        elif 'date' in filtered_data.columns:
            filtered_data['month'] = filtered_data['date'].dt.month
        else:
            filtered_data['month'] = np.nan
        
        # Define seasons (for India)
        def get_season(month):
            try:
                month = int(month)
                if month in [12, 1, 2]:
                    return 'Winter'
                elif month in [3, 4, 5]:
                    return 'Summer'
                elif month in [6, 7, 8, 9]:
                    return 'Monsoon'
                else:
                    return 'Post-Monsoon'
            except Exception:
                return 'Unknown'
        
        filtered_data['season'] = filtered_data['month'].apply(get_season)
        
        seasonal_data = filtered_data.groupby('season').agg({
            'Cases': 'mean',
            'Deaths': 'mean'
        }).reset_index()
        
        return jsonify({
            'seasons': seasonal_data['season'].tolist(),
            'avg_cases': [round(x, 1) for x in seasonal_data['Cases'].tolist()],
            'avg_deaths': [round(x, 1) for x in seasonal_data['Deaths'].tolist()]
        })
        
    except Exception as e:
        logger.error(f"Error in seasonal API: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/hospitals/search')
def api_hospital_search():
    """Search hospital data by city and year"""
    try:
        city = request.args.get('city', '')
        year = request.args.get('year', '')
        
        if not city or not year:
            return jsonify({'error': 'City and year are required'}), 400
        
        # Filter data for the specific city and year
        city_data = raw_data[
            (raw_data['city'].str.contains(city, case=False, na=False)) &
            (raw_data['year'] == int(year))
        ] if 'city' in raw_data.columns else pd.DataFrame()
        
        if city_data.empty:
            return jsonify({
                'found': False,
                'message': f'No data found for {city} in {year}'
            })
        
        # Aggregate results
        total_cases = int(city_data['Cases'].sum())
        total_deaths = int(city_data['Deaths'].sum())
        avg_temp = round(float(city_data['Temp'].mean()), 1) if 'Temp' in city_data.columns else None
        avg_preci = round(float(city_data['preci'].mean()), 2) if 'preci' in city_data.columns else None
        
        # Disease breakdown if available
        diseases = []
        if data_cache.get('disease_column') and data_cache['disease_column'] in city_data.columns:
            disease_col = data_cache['disease_column']
            disease_breakdown = city_data.groupby(disease_col)['Cases'].sum().to_dict()
            diseases = [{'name': disease, 'cases': int(cases)} for disease, cases in disease_breakdown.items()]
        else:
            # Mock disease breakdown
            diseases = [
                {'name': 'Dengue', 'cases': int(total_cases * 0.4)},
                {'name': 'Malaria', 'cases': int(total_cases * 0.3)},
                {'name': 'Typhoid', 'cases': int(total_cases * 0.2)},
                {'name': 'Others', 'cases': total_cases - int(total_cases * 0.9)}
            ]
        
        return jsonify({
            'found': True,
            'city': city,
            'year': year,
            'total_cases': total_cases,
            'total_deaths': total_deaths,
            'avg_temperature': avg_temp,
            'avg_precipitation': avg_preci,
            'diseases': diseases,
            'data_points': len(city_data)
        })
        
    except Exception as e:
        logger.error(f"Error in hospital search API: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/hospitals/list')
def api_hospital_list():
    """Get list of hospitals"""
    try:
        # Mock hospital data - replace with actual hospital database
        hospitals = [
            {
                'id': 1,
                'name': 'City General Hospital',
                'city': 'Mumbai',
                'state': 'Maharashtra',
                'beds': 500,
                'contact': '+91-22-1234-5678',
                'speciality': 'Infectious Diseases',
                'address': '123 Health Street, Mumbai'
            },
            {
                'id': 2,
                'name': 'Metro Medical Center',
                'city': 'Delhi',
                'state': 'Delhi',
                'beds': 750,
                'contact': '+91-11-9876-5432',
                'speciality': 'Emergency Care',
                'address': '456 Care Avenue, New Delhi'
            },
            {
                'id': 3,
                'name': 'Regional Health Institute',
                'city': 'Bangalore',
                'state': 'Karnataka',
                'beds': 400,
                'contact': '+91-80-5555-1111',
                'speciality': 'Tropical Diseases',
                'address': '789 Medical Plaza, Bangalore'
            },
            {
                'id': 4,
                'name': 'State Medical College',
                'city': 'Chennai',
                'state': 'Tamil Nadu',
                'beds': 600,
                'contact': '+91-44-2222-3333',
                'speciality': 'Public Health',
                'address': '321 University Road, Chennai'
            },
            {
                'id': 5,
                'name': 'District Hospital',
                'city': 'Pune',
                'state': 'Maharashtra',
                'beds': 350,
                'contact': '+91-20-7777-8888',
                'speciality': 'General Medicine',
                'address': '654 District Center, Pune'
            }
        ]
        
        city_filter = request.args.get('city', '').lower()
        if city_filter:
            hospitals = [h for h in hospitals if city_filter in h['city'].lower()]
        
        return jsonify({'hospitals': hospitals})
        
    except Exception as e:
        logger.error(f"Error in hospital list API: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/alerts/submit', methods=['POST'])
@login_required
def api_submit_alert():
    try:
        data = request.get_json()
        print("data is ", data)
        required_fields = ['name', 'email', 'phone', 'location', 'hospital_email', 'hospital_location', 'disease_type', 'cases', 'details']
        for field in required_fields:
            if field not in data or not data[field]:
                return jsonify({'error': f'Missing required field: {field}'}), 400

        conn = sqlite3.connect('alerts.db')
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO alerts (reporter_name, email, phone, location,hospital_email,hospital_location, disease, cases, details)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            data['name'], data['email'], data['phone'], data['location'],data['hospital_email'], data['hospital_location'],
            data['disease_type'], int(data['cases']), data.get('details', '')
        ))
        conn.commit()
        conn.close()

        return jsonify({
            'success': True,
            'message': 'Alert submitted successfully.',
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    
def api_submit_alert_old():
    """Submit disease outbreak alert"""
    try:
        data = request.get_json()
        
        required_fields = ['name', 'email', 'phone', 'location', 'disease_type', 'cases']
        for field in required_fields:
            if field not in data or not data[field]:
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        # Here you would typically save to a database
        alert_data = {
            'id': int(datetime.now().timestamp()),
            'timestamp': datetime.now().isoformat(),
            'name': data['name'],
            'email': data['email'],
            'phone': data['phone'],
            'location': data['location'],
            'disease_type': data['disease_type'],
            'cases': int(data['cases']),
            'details': data.get('details', ''),
            'status': 'submitted'
        }
        
        # Mock notification to medical centers
        logger.info(f"Alert submitted: {alert_data}")
        
        return jsonify({
            'success': True,
            'message': 'Alert submitted successfully. Nearest medical centers have been notified.',
            'alert_id': alert_data['id']
        })
        
    except Exception as e:
        logger.error(f"Error in submit alert API: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/data/metadata')

def api_metadata():
    """Get metadata about the dataset"""
    try:
        return jsonify(data_cache)
    except Exception as e:
        logger.error(f"Error in metadata API: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/data/debug')
@role_required(['admin', 'health_supervisor'])
def api_debug():
    """Debug endpoint to show what columns and diseases are available"""
    try:
        if raw_data is None:
            return jsonify({'error': 'No data loaded'})
        
        debug_info = {
            'columns': list(raw_data.columns),
            'shape': raw_data.shape,
            'disease_column': data_cache.get('disease_column'),
            'diseases': data_cache.get('diseases', []),
            'sample_data': {}
        }
        
        # Show sample disease data if available
        disease_col = data_cache.get('disease_column')
        if disease_col and disease_col in raw_data.columns:
            disease_counts = raw_data[disease_col].value_counts().head(10).to_dict()
            debug_info['disease_value_counts'] = disease_counts
            debug_info['sample_disease_rows'] = raw_data[disease_col].head(20).tolist()
        
        return jsonify(debug_info)
        
    except Exception as e:
        logger.error(f"Error in debug API: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/upload', methods=['POST'])
def upload_file():
    """Upload new CSV file"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if file and file.filename.endswith('.csv'):
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            
            # Create upload directory if it doesn't exist
            os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
            
            file.save(filepath)
            
            # Try to load the new file
            if load_and_process_data(filepath):
                return jsonify({'success': True, 'message': 'File uploaded and processed successfully'})
            else:
                return jsonify({'error': 'Failed to process the uploaded file'}), 400
        
        return jsonify({'error': 'Invalid file type. Only CSV files are allowed.'}), 400
        
    except Exception as e:
        logger.error(f"Error in file upload: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/show_alerts')
def show_alerts():
    conn = sqlite3.connect("alerts.db")
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM alerts")
    rows = cursor.fetchall()
    conn.close()
    return {"alerts": rows}

#################################################################################################################################################################

# ── Google Drive model loader ──────────────────────────────────────────────────
def download_from_gdrive(file_id: str, destination: str):
    """Download a file from Google Drive, handling the large-file confirm token."""
    URL = "https://docs.google.com/uc?export=download&confirm=t"
    session = requests.Session()
    response = session.get(URL, params={'id': file_id}, stream=True)

    # Extract confirmation token for files >100 MB
    token = next(
        (v for k, v in response.cookies.items() if k.startswith('download_warning')),
        None
    )
    if token:
        response = session.get(URL, params={'id': file_id, 'confirm': token}, stream=True)

    with open(destination, 'wb') as f:
        for chunk in response.iter_content(chunk_size=32768):
            if chunk:
                f.write(chunk)
    logger.info(f"Downloaded {destination} from Google Drive.")


MODEL_PATH   = 'disease_model.pkl'
COLUMNS_PATH = 'model_columns.pkl'

MODEL_GDRIVE_ID   = '1ZsUMNANPAwTqARS8RPRKKKhhYASL1wCM'   # disease_model.pkl
COLUMNS_GDRIVE_ID = 'YOUR_COLUMNS_FILE_ID_HERE'             # ← paste model_columns.pkl Drive file ID here

if not os.path.exists(MODEL_PATH):
    logger.info("Downloading disease_model.pkl from Google Drive ...")
    download_from_gdrive(MODEL_GDRIVE_ID, MODEL_PATH)
    logger.info("disease_model.pkl download complete.")

if not os.path.exists(COLUMNS_PATH):
    logger.info("Downloading model_columns.pkl from Google Drive ...")
    download_from_gdrive(COLUMNS_GDRIVE_ID, COLUMNS_PATH)
    logger.info("model_columns.pkl download complete.")

# Load model and columns
with open(MODEL_PATH, 'rb') as f:
    model = pickle.load(f)

with open(COLUMNS_PATH, 'rb') as f:
    model_columns = pickle.load(f)
# ──────────────────────────────────────────────────────────────────────────────

# Risk score function
def compute_risk_score(city, month, year, temp, preci, lai, latitude, longitude, past_cases=0):
    max_cases = 100  # Or any normalization constant
    past_score = float(past_cases)/max_cases if max_cases>0 else 0
    weather_score = min(1, (preci/10) + max(0, (temp-20)/20))
    season_score = 0.3 if month in [6,7,8,9] else 0
    risk_score = 0.5*past_score + 0.4*weather_score + 0.1*season_score
    risk_score = min(1, risk_score)
    if risk_score < 0.3:
        risk_level = 'Low'
    elif risk_score < 0.6:
        risk_level = 'Medium'
    else:
        risk_level = 'High'
    return risk_score, risk_level

# Prediction route
from flask import jsonify

@app.route('/predict', methods=['POST'])
@login_required
def predict():
    # Get data from form
    data = request.form
    input_dict = {
        'year':[int(data['year'])],
        'mon':[int(data['mon'])],
        'Cases':[float(data['Cases'])],
        'Deaths':[float(data['Deaths'])],
        'preci':[float(data['preci'])],
        'LAI':[float(data['LAI'])],
        'Temp_C':[float(data['Temp_C'])],
        'Latitude':[float(data['Latitude'])],
        'Longitude':[float(data['Longitude'])],
        'week_num':[int(data['week_num'])]
    }

    # Align columns
    for col in model_columns:
        if col not in input_dict:
            input_dict[col] = [0]

    input_df = pd.DataFrame(input_dict, columns=model_columns)

    # Predict disease probabilities
    proba = model.predict_proba(input_df)[0]
    disease_probs = pd.DataFrame({'Disease': model.classes_, 'Probability': proba}).sort_values(by='Probability', ascending=False)

    # Compute risk score
    risk_score, risk_level = compute_risk_score(
        city=data['city'],
        month=int(data['mon']),
        year=int(data['year']),
        temp=float(data['Temp_C']),
        preci=float(data['preci']),
        lai=float(data['LAI']),
        latitude=float(data['Latitude']),
        longitude=float(data['Longitude']),
        past_cases=float(data['Cases'])
    )

    # Send top 3 diseases + risk info to HTML
    top_diseases = disease_probs.head(3).to_dict(orient='records')

    return jsonify({
        'top_diseases': top_diseases,
        'risk_score': risk_score,
        'risk_level': risk_level
    })


if __name__ == '__main__':
    # Load initial data
    init_database() 
    if not load_and_process_data():
        logger.warning("Could not load initial data file. Dashboard will work with empty data.")
    
    # Create necessary directories
    os.makedirs('templates', exist_ok=True)
    os.makedirs('uploads', exist_ok=True)
    os.makedirs('downloads', exist_ok=True)
    
    app.run(debug=True, host='0.0.0.0', port=5000)